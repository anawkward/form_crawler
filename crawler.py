from bs4.element import *
from openpyxl import load_workbook
from pandas import DataFrame

# Baseline functions #############################################################################
def lowest_common_parent(soup, key1: str, key2: str):
    p1 = None
    p2 = None
    for t in soup.recursiveChildGenerator():
        if not isinstance(t, NavigableString):
            if t.text == key1:
                p1 = list(t.parents)[::-1]
            elif t.text == key2:
                p2 = list(t.parents)[::-1]
    if p1 is None or p2 is None:
        print('key not found !')
        return False
    common_parent = [x for x, y in zip(p1, p2) if x is y][-1]

    return common_parent

def get_element(node):
    ps = node.previous_siblings
    sib = []
    for t in ps:
        if t.name == node.name:
            sib.append(t)
    length = len(sib) + 1
    if length > 1:
        return '%s:nth-of-type(%s)' % (node.name, length)
    else:
        return node.name

def get_css_path(node):
    path = [get_element(node)]
    for parent in node.parents:
        if parent.name == 'body':
            break
        path.insert(0, get_element(parent))
    return ' > '.join(path)

def isdeepest(node:Tag):
    # deepestNode: a node which has no children or "only" NavigableString
    # so, deepestNode is either (1)NavigableString or (2)its children is empty
    if isinstance(node, NavigableString):
        return True
    if not any(True for _ in node.children): # if children is not not empty
        return True

def deepest_siblings(common_parent):
    retList = []
    for t in common_parent.recursiveChildGenerator():
        if isdeepest(t):
            retList.append(str(t.string))
    return retList

def list_to_excel(xlpath:str, sheet_num:int, dataList:list, ncol:int, byrow = True):
    wb = load_workbook(xlpath)
    ws = wb._sheets[sheet_num]
    r = ws.max_row
    if byrow is True:
        while dataList:
            for c in range(ncol):
                if not dataList:
                    break
                v = dataList.pop(0)
                wb._sheets[0].cell(row = 1+r, column = 1+c, value = v)
            r += 1
    else:
        print('not implemented yet')
        return
    wb.save(xlpath)
# Other useful functions #########################################################################
def children_to_texts(taglist:list, with_path = False):
    texts = ['']
    paths = []
    texpath = []
    for item in taglist:
        if isinstance(item, NavigableString):
            continue
        else:
            t = item.text
        if t:
            if texts[-1] != t:
                texts.append(t)
                paths.append(get_css_path(t))
            else:
                continue
    del texts[0]
    if with_path is False:
        return texts
    else:
        if with_path is True:
            for x, y in zip(texts, paths):
                texpath.append((x,y))
        return texts, paths, texpath

def remove_dup_rows(xlpath, outxlpath, sheet_name = "Sheet1"):
    wb = load_workbook(xlpath)
    if isinstance(sheet_name, int):
        sheet_name = wb.sheetnames[sheet_name]
    ws = wb[sheet_name]
    df = DataFrame(ws.values)
    df = df.drop_duplicates()
    df.to_excel(outxlpath, header = False, index = False)
# Key functions ######################################################################################
def crawl_first(soup, key1, key2):
    common_parent_tag = lowest_common_parent(soup, key1=key1, key2=key2)
    common_parent_path = get_css_path(common_parent_tag)
    deepest_siblings_list = deepest_siblings(common_parent_tag)
    length_between_keys = abs(deepest_siblings_list.index(key1)-deepest_siblings_list.index(key2))
    print(deepest_siblings_list[:length_between_keys])
    return common_parent_path, length_between_keys

def crawl_to_list(soup, common_parent_path):
    common_parent_tag = soup.select_one(common_parent_path)
    deepest_siblings_list = deepest_siblings(common_parent_tag)
    return deepest_siblings_list
######################################################################################################
'''
## example ##
import selenium.webdriver
import crawler
from bs4 import BeautifulSoup

option = selenium.webdriver.ChromeOptions()
driver = selenium.webdriver.Chrome(executable_path='C:/chromedriver.exe',options=option)  # if you don't have, get one on the internet
url = r"some website address you want to address XD"
driver.get(url)
##... manually navigate to the place you want to scrape ... ##
# keys are to find "table where?" and "how many columns?" by analyzing distance between two keys with one row difference.
key1 = 'book_name_in_a_first_row'
key2 = 'book_name_in_a_second_row'
# or key2 = 'another_column_in_a_first_row_if_it_is_one_liner'

page_source = driver.page_source
soup = BeautifulSoup(page_source, 'html.parser')
common_parent_path, length_between_keys = crawler.crawl_first(soup, key1, key2)

xlpath = './myscraping.xlsx'
sheet_num = 0
page_source = driver.page_source
soup = BeautifulSoup(page_source, 'html.parser')
ncol = length_between_keys

# check dataList, if it's good.
dataList = crawler.crawl_to_list(soup, common_parent_path)[0:]
# if you have an empty excel file, save texts there.
# if excel file isn't empty, data is added along rows
crawler.list_to_excel(xlpath, sheet_num, dataList, ncol)
'''